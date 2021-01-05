from openpyxl import Workbook, load_workbook

from .report_lib import *

from .models import Person, Query, Report, Exception, Url, Url_data

from django.shortcuts import redirect
import datetime

def save_data_for_report(request, person_id):

    person = Person.objects.get(pk=person_id)
    person.status = 1
    person.save()

    person.mkdir()

    queries = Query.objects.filter(person=person)

    driver = webdriver.Chrome(options=set_options())
    driver.maximize_window()

    save_data_from_search_system(driver=driver, queries=queries, search_system='Yandex', person=person)
    save_data_from_search_system(driver=driver, queries=queries, search_system='Google', person=person)

    person.status = 2
    person.save()

    driver.close()
    response = redirect('/')
    return response

def create_report_table(request, person_id):
    person = Person.objects.get(pk=person_id)

    screenshots_folder_name = f'Person_{person.id}/{person.last_it}/screenshots/'

    queries = Query.objects.filter(person=person)

    mark_images_as_negative(queries, person.last_it)

    template_wb = load_workbook(f'{person.start_table}')
    ws_din = template_wb['Динамика по запросам']

    current_date = datetime.date.today()
    if ws_din[3][-1].value.date() != current_date:
        add_date(ws_din, len(ws_din[3]), len(ws_din['B']), len(queries), current_date)

    count_neg_urls(ws_din, queries, person.last_it)

    template_wb.save(f'{person.start_table}')

    result_wb = load_workbook(f'{person.start_table}')

    ws_google_screenshots = result_wb.create_sheet('Скриншоты Google')
    ws_yandex_screenshots = result_wb.create_sheet('Скриншоты Yandex')
    ws_urls = result_wb.create_sheet('Выдача')


    indent_yandex = 0
    indent_google = 0
    indent_result = 1

    for query in queries:
        save_screenshots_for_query(query=query, last_it=person.last_it, search_system='Yandex', folder=screenshots_folder_name)
        save_screenshots_for_query(query=query, last_it=person.last_it, search_system='Google', folder=screenshots_folder_name)

        indent_yandex = add_screenshots_to_ws(ws_yandex_screenshots, f'{screenshots_folder_name}{query.id}_yandex_first_page.png',
                                              f'{screenshots_folder_name}{query.id}_yandex_second_page.png', indent_yandex)
        indent_google = add_screenshots_to_ws(ws_google_screenshots, f'{screenshots_folder_name}{query.id}_google_first_page.png',
                                              f'{screenshots_folder_name}{query.id}_google_second_page.png', indent_google)

        indent_result = add_results_to_ws(ws=ws_urls, query=query, last_it=person.last_it, indent=indent_result)

    report = Report(person=person, last_it=person.last_it)
    report.table_name = f'Person_{person.id}/{person.last_it}/tables/{person.last_it}_{person.pk}_отчет_{person}_{current_date}.xlsx'
    report.table = report.table_name
    report.last_it = person.last_it
    report.save()


    result_wb.save(f'{report.table_name}')

    person.mkit()

    response = waprfile(f'{report.table_name}', f'отчет_{str(person)}_{current_date}')
    return response


def drop_status(request, person_id):
    person = Person.objects.get(pk=person_id)
    person.status = 0
    person.save()
    return redirect('/')
