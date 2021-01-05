import os
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import random

from selenium.webdriver.remote.webelement import WebElement

import requests

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, colors

from copy import copy
import datetime

from .ru_captcha_python_3_4 import ansver

from PIL import ImageOps, Image

from django.http import HttpResponse

from urllib.parse import quote

from .models import Url, Url_data, Query, Person, Exception, Report

import creds

footers_path = 'report/static/footers'

def get_footers_path(system):
    return f'{footers_path}/{system}'

def save_element_screenshot(element: WebElement, name: str, scroll=1):
    """
    Save screenshot of WebElement and scroll page
    """

    element.screenshot(f'{name}.png')
    if (scroll):
        driver = element._parent
        size = element.size
        h_elem = size['height']
        script = f'window.scrollTo(0, window.scrollY + {h_elem + 10})'
        driver.execute_script(script)

def elem_process(elem, exceptions, name):
    """
    Check if element in exceptions or ads
    return url if there were no errors
    else return 'Error'
    """

    href = elem.find_element_by_tag_name('a').get_attribute('href')
    if href not in exceptions and href.find('youtube.com') == -1 and href.find('yabs.yandex.ru') == -1:
        save_element_screenshot(elem, name)
        return href
    return 'Error'

def google_search(driver, req):
    """
    put request into google search field
    and press find
    """

    elem = driver.find_element_by_name("q")
    elem.clear()
    elem.send_keys(req)
    time.sleep(random.uniform(1, 3))
    elem.send_keys(Keys.RETURN)

def yandcaptcha(driver):
    """
    looking for yandex captcha, download image and solve captcha
    """

    try:
        driver.find_element_by_xpath("//div[@class='captcha__image']")
        src = driver.find_element_by_xpath("//html//body//div//form//div[1]//div//div[1]//img").get_attribute('src')
        print(src)
        p = requests.get(src)
        out = open("image.png", "wb")
        out.write(p.content)
        out.close()
        anns = ansver()
        elem = driver.find_element_by_xpath("//input[@name='rep']")
        elem.send_keys(anns)
        elem.send_keys(Keys.RETURN)
    except:
        pass

def set_options():
    """
    set optionts to driver
    """

    options = webdriver.ChromeOptions()
    options.add_argument("--incognito")
    '''options.add_argument('--no-sandbox')
    options.add_argument('--headless')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("--disable-gpu")'''
    return options

def yandex_search_settings(driver):
    """
    set yandex search settings
    """

    driver.get('https://yandex.ru/tune/search')
    driver.find_element_by_xpath('//*[@id="form__a11y"]/div[2]/div[2]/div[1]/span/label').click()
    driver.find_element_by_xpath('//*[@id="form__a11y"]/div[2]/div[3]/div[1]/span/label').click()
    driver.find_element_by_xpath('//*[@id="form__a11y"]/div[2]/div[4]/div[1]/span/label').click()
    driver.find_element_by_xpath('//*[@id="form__a11y"]/div[3]/div/button').click()
    time.sleep(1)
    driver.get('https://yandex.ru')

def google_search_settings(driver):
    """
    set google search settings
    """

    driver.get('https://www.google.com/history/optout?hl=ru')
    driver.find_element_by_xpath("//input[@type='checkbox']").click()
    driver.get('https://google.com')

def yandex_search(driver, request):
    """
    put reguest into yandex search field an press search
    """

    elem = driver.find_element_by_class_name("input__control,input__input")  # ищет куда бы ему запрос ввести
    elem.clear()
    elem.send_keys(request)
    time.sleep(1)
    elem.send_keys(Keys.RETURN)

def captcha_alert(driver):
    """
    find google captcha and solve it
    """

    status = 0
    try:
        driver.find_element_by_id('recaptcha')
        status = 1
        strok = driver.find_element_by_xpath('//*[@id="recaptcha"]')
        googlekey = str(strok.get_attribute('data-sitekey'))
        datas = strok.get_attribute('data-s')
        key = creds.CAPTCHA_API_KEY
        method = 'userrecaptcha'
        site = str(driver.current_url)
        exp = requests.get(
            'https://rucaptcha.com/in.php?key=' + key + '&method=' + method + '&googlekey=' + googlekey + '&pageurl=' + site + '&data-s=' + datas)
        id_cap = str(exp.text).split('|')[1]
        time.sleep(25)
        exp2 = requests.get('https://rucaptcha.com/res.php?key=' + key + '&action=get&id=' + id_cap)
        test_text = str(exp2.text)
        while test_text == 'CAPCHA_NOT_READY':
            time.sleep(10)
            exp2 = requests.get('https://rucaptcha.com/res.php?key=' + key + '&action=get&id=' + id_cap)
            test_text = str(exp2.text)
        id_cap2 = str(exp2.text).split('|')[1]
        elem = driver.find_element_by_id('g-recaptcha-response')
        driver.execute_script("arguments[0].style.display = '';", elem)
        elem.send_keys(id_cap2)
        elem2 = driver.find_element_by_xpath('//*[@id="captcha-form"]/script[3]')
        driver.execute_script("submitCallback()", elem2)
        status = 0
    except:
        pass
    return status

def get_merged_screenshots(urls):
    """
    merge screenshots in Url list
    """

    images = [Image.open(e) for e in urls]
    sizes = [e.size for e in images]
    widths, heights = zip(*sizes)
    screen_height = sum(heights) + 20 * len(sizes)

    img = Image.new('RGB', (widths[1] + 10, screen_height), color='white')

    h = 0
    for im in images:
        img.paste(im, (0, h))
        h += 20 + im.size[1]
    return img


def mark_images_as_negative(queries, last_it):
    """
    add red boards to all negative urls in queries
    """

    for query in queries:
        urls = Url.objects.filter(query=query, last_it=last_it)
        for url in urls:
            if url.url_data.tone:
                ImageOps.expand(Image.open(f'{url.image}.png'), border=5, fill='red').save(f'{url.image}.png')


def save_screenshots_for_query(query, last_it, search_system, folder=''):
    """
    create 2 merged screenshots of search system for query
    """

    urls = Url.objects.filter(query=query, last_it=last_it, search_system=search_system)[:20]

    system = search_system.lower()

    if system == 'google':
        header = query.google_header
    else:
        header = query.yandex_header

    elems = [header]
    elems += [f'{e.image}.png' for e in urls[:10]]
    elems.append(f'{get_footers_path(system)}/{system}_footer_1.png')
    first_page = get_merged_screenshots(elems)

    elems = [header]
    elems += [f'{e.image}.png' for e in urls[10:]]
    elems.append(f'{get_footers_path(system)}/{system}_footer_2.png')
    second_page = get_merged_screenshots(elems)

    first_page.save(f'{folder}{query.id}_{search_system}_first_page.png')
    second_page.save(f'{folder}{query.id}_{search_system}_second_page.png')

def add_screenshots_to_ws(ws, screen_1, screen_2, indent):
    """
    add screenshots to excel worksheet
    """

    my_png = openpyxl.drawing.image.Image(screen_1)
    cell = 'A' + str(2 + indent)
    ws.add_image(my_png, cell)
    my_png = openpyxl.drawing.image.Image(screen_2)
    cell = 'O' + str(2 + indent)
    ws.add_image(my_png, cell)
    max_height = max(Image.open(screen_1).size[1], Image.open(screen_2).size[1])
    indent += int(max_height/16)
    return indent

def add_results_to_ws(ws, query, last_it, indent):
    """
    add search result to worksheet
    """

    urls_yandex = Url.objects.filter(query=query, last_it=last_it, search_system='Yandex')[:20]
    urls_google = Url.objects.filter(query=query, last_it=last_it, search_system='Google')[:20]


    ws.cell(row=1, column=indent+1, value=str(query))
    ws.cell(row=1, column=indent + 3, value=' ')
    ws.merge_cells(f'{get_column_letter(indent+1)}1:{get_column_letter(indent+2)}1')
    ws.cell(row=2, column=indent+1, value='Яндекс')
    ws.cell(row=2, column=indent+2, value='Google')
    print(query.pk, len(urls_yandex))
    for i in range(20):
        print(i)
        print(urls_yandex[i])
        if urls_yandex[i].url_data.tone:
            ws.cell(row=i + 3, column=indent + 1).font = Font(color="FF0000", italic=True, bold=True)
        ws.cell(row=i+3, column=indent+1, value=urls_yandex[i].url_data.href)
        if urls_google[i].url_data.tone:
            ws.cell(row=i + 3, column=indent + 2).font = Font(color="FF0000", italic=True, bold=True)
        ws.cell(row=i+3, column=indent + 2, value=urls_google[i].url_data.href)
        ws.cell(row=i+3, column=indent + 3, value=' ')
    return indent+4


def copy_cell_fomat(cell, new_cell):
    """
    copy format of Excel cell
    """

    if cell.has_style:
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)


def copy_column(ws, last_column, column_len):
    """
    copy format of last column in worksheet
    into right collumn
    """

    for i in range(column_len):
        cell = ws.cell(row=i+1, column=last_column)
        new_cell = ws.cell(row=i+1, column=last_column+1)
        copy_cell_fomat(cell, new_cell)

def add_date(ws, last_column, column_len, number_of_queries, date):
    """
    add new date in report.xlsx
    """

    copy_column(ws, last_column, column_len)
    ws.column_dimensions[get_column_letter(last_column + 1)].width = 3.86
    current_date = datetime.date.today()
    ws.cell(row=3, column=last_column+1, value=current_date)
    ws.cell(row=6+number_of_queries, column=last_column+1, value=date)
    ws[3][last_column-1].fill = PatternFill("solid", fgColor=colors.BLACK)
    ws[6+number_of_queries][last_column-1].fill = PatternFill("solid", fgColor=colors.BLACK)
    for i in range(4, 4+number_of_queries):
        k = i + 3 + number_of_queries
        ws[i][last_column-1].fill = PatternFill("solid", fgColor=colors.WHITE)
        ws[i][last_column-1].font = Font(name='Calibri', size=8, bold=False, color=colors.BLACK)
        ws[k][last_column-1].fill = PatternFill("solid", fgColor=colors.WHITE)
        ws[k][last_column-1].font = Font(name='Calibri', size=8, bold=False, color=colors.BLACK)

def count_neg_urls(ws, queries, last_it):
    """
    count neg urls for every query and search system
    then add it to worksheet
    """

    n = 1
    ws[4+len(queries)][-1].value = f'=SUM({get_column_letter(len(ws[3]))}{4}:{get_column_letter(len(ws[3]))}{3+len(queries)})'
    ws[7 + 2*len(queries)][-1].value = f'=SUM({get_column_letter(len(ws[3]))}{7+len(queries)}:{get_column_letter(len(ws[3]))}{6 + 2*len(queries)})'
    ws[8 + 2*len(queries)][-1].value = f'={get_column_letter(len(ws[3]))}{4+len(queries)}+{get_column_letter(len(ws[3]))}{7 + 2*len(queries)}'
    ws[100][-1].value = 200
    ws[100][-2].value = 0
    for query in queries:
        urls_google = Url.objects.filter(query=query, last_it=last_it, search_system='Google')[:20]
        urls_yandex = Url.objects.filter(query=query, last_it=last_it, search_system='Yandex')[:20]
        google_sum = sum([e.url_data.tone for e in urls_google])
        yandex_sum = sum([e.url_data.tone for e in urls_yandex])
        ws[3+n][-1].value = google_sum
        ws[6 + len(queries) + n][-1].value = yandex_sum
        n += 1


def waprfile(file, name):
    """
    send excel file to user
    """

    with open(file, "rb") as excel:
        data = excel.read()
    response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename*=utf-8''{quote(name)}.xlsx"
    return response



def captcha_alert_for_sr(driver, sr):
    """
    choose right way to solve captcha for curent search system
    """

    if sr == 'Google':
        while captcha_alert(driver):
            pass
    else:
        yandcaptcha(driver)

def search(driver, req, sr):
    """
    choose right way to pul request into search system field
    """

    if sr == 'Google':
        google_search(driver, req)
    else:
        yandex_search(driver, req)

def search_settings(driver, sr):
    """
    Choose right way to set settings for current search system
    """

    if sr == 'Google':
        google_search_settings(driver)
    else:
        yandex_search_settings(driver)

def save_data_from_search_system(driver, search_system, person, queries, exceptions=[]):

    """
    get all urls and screenshots from current search system
    also save all of them in database
    """

    screenshots_folder_name = person.get_screenshots_path

    search_settings(driver, search_system)
    captcha_alert_for_sr(driver, search_system)

    if search_system == 'Yandex':
        elem_xpath = "//div[@class='serp-header__main']"
        search_result_xpath = "//ul[@id='search-result']/li"
        next_page = "//a[@aria-label='Следующая страница']"
    else:
        elem_xpath = "//*[@id='searchform']"
        search_result_xpath = "//div[@id='rso']//div[@class='g' or @style='box-shadow:none' or @style='position:relative']"
        next_page = "//*[@id='pnnext']"

    sr = search_system.lower()

    for query in queries:
        search(driver, query.query, search_system)
        captcha_alert_for_sr(driver, search_system)
        elem = driver.find_element_by_xpath(elem_xpath)
        save_element_screenshot(element=elem, name=f'Person_{person.id}/{query.id}_{sr}_header', scroll=0)

        if sr == 'yandex':
            query.yandex_header = f'Person_{person.id}/{query.id}_{sr}_header.png'
        else:
            query.google_header = f'Person_{person.id}/{query.id}_{sr}_header.png'

        query.save()
        position = 0
        error_req = 0
        for page_number in range(1, 4):
            page_position = 0
            error_page = 0
            search_result = driver.find_elements_by_xpath(search_result_xpath)
            for elem in search_result:
                position += 1
                page_position += 1
                image_name = f'{screenshots_folder_name}/{query.pk}_{position}_{page_number}_{sr}'
                href = elem_process(elem, exceptions, image_name)
                if href != 'Error':
                    lst = Url_data.objects.filter(href=href, person=person)
                    if len(lst):
                        url_data = lst[0]
                    else:
                        url_data = Url_data(href=href, person=person)
                        url_data.save()
                    url = Url(query=query, position=position - error_req, page_number=page_number,
                              page_position=page_position - error_page, search_system=search_system,
                              image=image_name, href=href, last_it=person.last_it, url_data=url_data)
                    url.save()
                else:
                    error_req += 1
                    error_page += 1
            driver.find_element_by_xpath(next_page).click()

